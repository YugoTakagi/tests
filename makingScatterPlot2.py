import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
import pandas as pd
import os
from tqdm import tqdm
################################################################################
# inputFilePath = [str(input("Please write FilePath(~.csv)>>> "))]
inputFilePath = input("Please write FilePath(~.csv,~.csv,..)>>> ")
IFP = inputFilePath.split(",")

for iP in tqdm(IFP):
    data = pd.read_csv(iP)
    ######
    oF = iP.rsplit(".",1)
    ######
    outputFilePath = "{}.xlsx".format(oF[0])
    data.to_excel(outputFilePath, encoding='utf-8', index=False, header=False)
    # print('outputFilePath = {}'.format(outputFilePath))

    ################################################################################
    # Excelファイルを開きます
    # outputFilePath = str(input("Please write FilePath(~.xlsx)>>> ")) # 出力したいExcelファイル名を指定してください
    wb = openpyxl.load_workbook(outputFilePath)
    sheetNames = wb.sheetnames
    # print("--  sheetNames = {}  --".format(sheetNames))
    ws = wb[sheetNames[0]]  # ここでは例として1つ目のシートを取得しています


    # 散布図を用意します
    cht = openpyxl.chart.ScatterChart()
    cht.title = '加速度センサの実験'
    cht.x_axis.title = 'num'
    cht.y_axis.title = "Arudino出力"
    cht.y_axis.majorGridLines = None
    cht.height = 12
    cht.width = 24
    # グラフ化するデータを参照する
    startRow = 1
    endRow = 0
    ################################################################################
    for cell_obj in list(ws.columns)[0]:#list()[n]  #n :=  a=0, b=1, ...
        endRow += 1
    ################################################################################

    # Excelのデータ参照した変数を用意する
    numValues = openpyxl.chart.Reference(ws, min_col=1, min_row=startRow, max_row=endRow)
    xValues = openpyxl.chart.Reference(ws, min_col=2, min_row=startRow, max_row=endRow)
    yValues = openpyxl.chart.Reference(ws, min_col=3, min_row=startRow, max_row=endRow)
    zValues = openpyxl.chart.Reference(ws, min_col=4, min_row=startRow, max_row=endRow)
    # print("xValues = {}".format(xValues))
    # print("yValues = {}".format(yValues))
    # print("zValues = {}".format(zValues))

    # 系列を用意し、データ参照を入力する
    s1 = openpyxl.chart.Series(xValues, numValues, title_from_data=False, title="x")
    s2 = openpyxl.chart.Series(yValues, numValues, title_from_data=False, title="y")
    s3 = openpyxl.chart.Series(zValues, numValues, title_from_data=False, title="z")

    # グラフの書式設定をする
    s1.graphicalProperties.line.solidFill = "4f81bd" # グラフの線の色
    s1.marker.symbol = "diamond"# 各データ点のマーカーの形状
    s1.marker.graphicalProperties.solidFill = "4f81bd" # 各データ点のマーカーの塗りつぶし色
    s1.marker.graphicalProperties.line.solidFill = "4f81bd"# 各データ点のマーカーの枠の色

    s2.graphicalProperties.line.solidFill = "8064a2"
    s2.marker.symbol = "triangle"
    s2.marker.graphicalProperties.solidFill = "8064a2"
    s2.marker.graphicalProperties.line.solidFill = "8064a2"

    s3.graphicalProperties.line.solidFill = "9bbb59"
    s3.marker.symbol = "triangle"
    s3.marker.graphicalProperties.solidFill = "9bbb59"
    s3.marker.graphicalProperties.line.solidFill = "9bbb59"


    # Chartに系列を追加する
    cht.series.append(s1)
    cht.series.append(s2)
    cht.series.append(s3)


    # Excelシートにグラフを貼り付ける
    graphInsertCol = 2 # グラフを挿入する列番号
    inColLetter = get_column_letter(graphInsertCol)
    inRow = 4 # グラフを挿入する行番号
    inCellLetter = inColLetter + str(inRow) # グラフを挿入するセルの位置をExcel形式で作る
    ws.add_chart(cht, inCellLetter)

    ######
    ws['G1'] = 'x'
    ws['G2'] = '=sum(B{}:B{})/{}'.format(startRow, endRow, endRow)
    ws['H1'] = 'y'
    ws['H2'] = '=sum(C{}:C{})/{}'.format(startRow, endRow, endRow)
    ws['I1'] = 'z'
    ws['I2'] = '=sum(D{}:D{})/{}'.format(startRow, endRow, endRow)
    ######
    oF = outputFilePath.rsplit(".",1)
    wb.save("{}'.xlsx".format(oF[0]))
    # print('>outputFilePath = {}'.format(oF))
    os.remove(outputFilePath)
print("--> success")
